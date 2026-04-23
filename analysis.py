import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart

print(" Starting Data Processing...")

# Load data
df = pd.read_csv("online_retail.csv", encoding='ISO-8859-1')

# Cleaning
df = df.dropna().drop_duplicates()
df['InvoiceDate'] = pd.to_datetime(df['InvoiceDate'])
df = df[(df['Quantity'] > 0) & (df['UnitPrice'] > 0)]

# Features
df['TotalSales'] = df['Quantity'] * df['UnitPrice']
df['Month'] = df['InvoiceDate'].dt.month

print(" Data Ready")

# Save to SQL
conn = sqlite3.connect("database.db")
df.to_sql("sales", conn, if_exists="replace", index=False)

print(" SQL Database Created")

# Excel Dashboard
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Dashboard"

# ---------------------------
#  Custom Styles (Unique)
# ---------------------------
title_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
card_fill = PatternFill(start_color="F3F6FF", end_color="F3F6FF", fill_type="solid")

title_font = Font(color="FFFFFF", bold=True, size=16)
label_font = Font(bold=True, size=12)
value_font = Font(bold=True, size=13)

center = Alignment(horizontal="center")

# ---------------------------
#  Title
# ---------------------------
ws.merge_cells("A1:E1")
ws["A1"] = " Retail Sales Insight Dashboard"
ws["A1"].fill = title_fill
ws["A1"].font = title_font
ws["A1"].alignment = center

# ---------------------------
#  KPI Calculations
# ---------------------------
total_sales = df['TotalSales'].sum()
transactions = df['InvoiceNo'].nunique()
countries = df['Country'].nunique()
products = df['Description'].nunique()

# ---------------------------
#  KPI Cards Layout
# ---------------------------
cards = [
    ("Total Revenue", round(total_sales)),
    ("Transactions", transactions),
    ("Countries", countries),
    ("Products", products),
]

row = 3
col = 1

for title, value in cards:
    ws.cell(row=row, column=col, value=title).font = label_font
    ws.cell(row=row+1, column=col, value=value).font = value_font

    ws.cell(row=row, column=col).fill = card_fill
    ws.cell(row=row+1, column=col).fill = card_fill

    ws.cell(row=row, column=col).alignment = center
    ws.cell(row=row+1, column=col).alignment = center

    col += 2  # spacing between cards

# ---------------------------
#  Insights Section
# ---------------------------
ws["A7"] = " Key Insights"
ws["A7"].font = label_font

top_country = df.groupby('Country')['TotalSales'].sum().idxmax()
top_product = df.groupby('Description')['TotalSales'].sum().idxmax()

ws["A9"] = "Top Country"
ws["B9"] = top_country

ws["A10"] = "Top Product"
ws["B10"] = top_product

# ---------------------------
#  Column spacing
# ---------------------------
for col in ["A", "B", "C", "D", "E"]:
    ws.column_dimensions[col].width = 22

# Save
wb.save("automated_dashboard.xlsx")

print("Excel Dashboard Created!")

# Country Sales
country_df = df.groupby('Country')['TotalSales'].sum().reset_index()
ws2 = wb.create_sheet("Country Sales")
ws2.append(["Country", "TotalSales"])
for row in country_df.values:
    ws2.append(list(row))

chart = BarChart()
data = Reference(ws2, min_col=2, min_row=1, max_row=len(country_df)+1)
cats = Reference(ws2, min_col=1, min_row=2, max_row=len(country_df)+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.title = "Sales by Country"
ws2.add_chart(chart, "E5")

# Monthly Sales
monthly_df = df.groupby('Month')['TotalSales'].sum().reset_index()
ws3 = wb.create_sheet("Monthly Sales")
ws3.append(["Month", "TotalSales"])
for row in monthly_df.values:
    ws3.append(list(row))

line = LineChart()
data = Reference(ws3, min_col=2, min_row=1, max_row=len(monthly_df)+1)
cats = Reference(ws3, min_col=1, min_row=2, max_row=len(monthly_df)+1)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.title = "Monthly Sales"
ws3.add_chart(line, "E5")

# Save Excel
wb.save("automated_dashboard.xlsx")

print(" Excel Dashboard Created!")